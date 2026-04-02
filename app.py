import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import warnings
warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & GLOBAL STYLES
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Flipkart Reconciliation – Yash Gallery",
    layout="wide",
    page_icon="🧾",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
/* ── Typography & Base ── */
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }

/* ── App background ── */
.stApp { background: #f5f7fa; }
.block-container { padding: 1.4rem 2rem 2rem; max-width: 1600px; }

/* ── Sidebar background ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f2139 0%, #1a3c5e 100%);
    border-right: 2px solid #2563a8;
}

/* Sidebar plain text & markdown */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] li,
[data-testid="stSidebar"] span:not([data-testid]),
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] .stMarkdown li,
[data-testid="stSidebar"] [data-testid="stCaption"] {
    color: #e8f0f9 !important;
}

/* Sidebar headings */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] .stMarkdown h1,
[data-testid="stSidebar"] .stMarkdown h2,
[data-testid="stSidebar"] .stMarkdown h3 { color: #ffd700 !important; }

/* Sidebar widget labels */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stFileUploader label,
[data-testid="stSidebar"] .stNumberInput label {
    color: #93c5fd !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
}

/* File uploader box — keep white background so drag text is readable */
[data-testid="stSidebar"] [data-testid="stFileUploader"] section {
    background: #ffffff !important;
    border: 1.5px dashed #93c5fd !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploader"] section * {
    color: #0f2139 !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploader"] section button {
    background: #1d4ed8 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 6px !important;
}

/* Number inputs — white box with dark text */
[data-testid="stSidebar"] input[type="number"] {
    background: #ffffff !important;
    color: #0f2139 !important;
    border: 1px solid #93c5fd !important;
    border-radius: 6px !important;
}

/* HR divider */
[data-testid="stSidebar"] hr { border-color: #2563a8; }

/* Success / info messages inside sidebar */
[data-testid="stSidebar"] .stAlert { color: #0f2139 !important; }

/* ── Metric cards ── */
[data-testid="stMetric"] {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 0.7rem 1rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
[data-testid="stMetricValue"] { font-size: 1.15rem !important; font-weight: 700 !important; color: #0f2139 !important; }
[data-testid="stMetricLabel"] { font-size: 0.72rem !important; color: #64748b !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.04em; }
[data-testid="stMetricDelta"] { font-size: 0.78rem !important; font-weight: 600 !important; }

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"] { border-bottom: 2px solid #dbeafe; gap: 0.25rem; }
[data-testid="stTabs"] button[role="tab"] {
    font-weight: 600; font-size: 0.85rem; color: #64748b;
    border-radius: 8px 8px 0 0; padding: 0.55rem 1.2rem;
    border: none; background: transparent; transition: all .2s;
}
[data-testid="stTabs"] button[role="tab"]:hover { background: #eff6ff; color: #1d4ed8; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background: #1d4ed8; color: #fff !important;
    box-shadow: 0 -2px 6px rgba(29,78,216,0.25);
}

/* ── Expanders ── */
[data-testid="stExpander"] { border: 1px solid #dbeafe; border-radius: 10px; background: #fff; }

/* ── Buttons ── */
.stButton button {
    border-radius: 8px; font-weight: 600; font-size: 0.85rem;
    transition: all .2s; border: none;
}
.stButton button[kind="primary"] {
    background: linear-gradient(135deg, #1d4ed8, #2563eb);
    color: #fff; box-shadow: 0 2px 6px rgba(29,78,216,0.3);
}
.stButton button[kind="primary"]:hover { box-shadow: 0 4px 12px rgba(29,78,216,0.45); transform: translateY(-1px); }

/* ── Download buttons ── */
.stDownloadButton button {
    background: linear-gradient(135deg, #0f766e, #14b8a6);
    color: #fff; border-radius: 8px; font-weight: 600;
    box-shadow: 0 2px 6px rgba(20,184,166,0.3); border: none;
}
.stDownloadButton button:hover { box-shadow: 0 4px 12px rgba(20,184,166,0.45); transform: translateY(-1px); }

/* ── DataFrames ── */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; border: 1px solid #e2e8f0; }

/* ── Info / success / warning boxes ── */
.stAlert { border-radius: 10px; font-size: 0.88rem; }

/* ── Section headers ── */
.section-header {
    background: linear-gradient(135deg, #0f2139, #1e3a5f);
    color: #ffd700; padding: 0.55rem 1rem; border-radius: 8px;
    font-weight: 700; font-size: 0.95rem; margin: 0.8rem 0 0.6rem;
    letter-spacing: 0.02em;
}

/* ── Brand badges ── */
.brand-badge {
    display: inline-block; padding: 2px 8px; border-radius: 20px;
    font-size: 0.72rem; font-weight: 700; letter-spacing: 0.04em;
    text-transform: uppercase;
}

/* ── Hero banner ── */
.hero-banner {
    background: linear-gradient(135deg, #0f2139 0%, #1a3c5e 60%, #1d4ed8 100%);
    border-radius: 14px; padding: 1.4rem 2rem;
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 1.2rem; box-shadow: 0 4px 18px rgba(15,33,57,0.25);
}
.hero-title { color: #ffd700; font-size: 1.6rem; font-weight: 800; margin: 0; letter-spacing: -0.01em; }
.hero-sub   { color: #93c5fd; font-size: 0.8rem; margin: 0.2rem 0 0; font-weight: 500; }
.hero-badge { background: rgba(255,215,0,0.15); border: 1px solid #ffd700;
               color: #ffd700; border-radius: 8px; padding: 0.35rem 0.85rem;
               font-size: 0.75rem; font-weight: 700; }

/* ── Progress / spinner ── */
.stSpinner { color: #1d4ed8; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HERO BANNER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero-banner">
  <div>
    <p class="hero-title">🧾 Flipkart Reconciliation</p>
    <p class="hero-sub">Yash Gallery Private Limited &nbsp;·&nbsp; Finance Team &nbsp;·&nbsp; Built by Ashu Bhatt</p>
  </div>
  <div class="hero-badge">v2.0 · Enhanced</div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
TDS_RATE = 0.001
TCS_RATE = 0.005

MONEY_COLS = [
    "Invoice Amount", "GT (As Per Calc)", "Selling Price", "Commission",
    "Collection Fee", "Fixed Fee", "Total Charges", "GST on Charges",
    "Taxable Value", "TDS", "TCS", "Total Deductions", "Received Amount",
    "PWN", "PWN Benchmark", "Difference",
]

REQUIRED_ORDER_COLS = {"Order Id", "SKU", "Invoice Amount", "Quantity", "Product"}

# ── Brand detection ──────────────────────────────────────────────────────────
KNOWN_BRANDS = [
    "Yash Gallery",
    "HouseOfCommon",
    "KALINI",
    "Tasrika",
    "AKIKO",
    "IKRASS",
]
BRAND_DISPLAY_MAP = {
    "AKIKO":         "Pushpa",
    "HouseOfCommon": "Pushpa",
}

# ── Size expand ─────────────────────────────────────────────────────────────
SIZE_EXPAND = {
    "L-XL":["L","XL"], "S-M":["S","M"], "XXL-3XL":["XXL","3XL"],
    "F-S/XXL":["F"], "F-3xl/5xl":["F"], "XS-S":["XS","S"],
    "M-L":["M","L"], "XL-XXL":["XL","XXL"], "3XL-4XL":["3XL","4XL"],
    "5XL-6XL":["5XL","6XL"], "7XL-8XL":["7XL","8XL"],
    "4XL-5XL":["4XL","5XL"], "2XL-3XL":["2XL","3XL"],
    "XS-S-M":["XS","S","M"], "L-XL-XXL":["L","XL","XXL"],
}
ORDER_TO_PRICE_SIZE: dict = defaultdict(list)
for _ps, _os_list in SIZE_EXPAND.items():
    for _os in _os_list:
        ORDER_TO_PRICE_SIZE[_os.upper()].append(_ps)

VENDOR_PREFIXES = ["GWN-","GWN_","GWN","SPF-","SPF_","SPF","KL_","KL-","KL"]

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 📂 Upload Files")
    order_files = st.file_uploader(
        "1️⃣  Order File(s)  (CSV / XLSX / XLS — multiple allowed)",
        type=["csv","xlsx","xls"], accept_multiple_files=True,
        help="Flipkart Seller Hub export. Can upload multiple files — they will be merged."
    )
    charges_file = st.file_uploader(
        "2️⃣  Data Excel",
        type=["xlsx"],
        help="Sheet 0: Charge Rates  |  Sheet 1: SKU→Sub-Category  |  Sheet 2: PWN Prices"
    )
    replace_sku_file = st.file_uploader(
        "3️⃣  Replace SKU Excel (optional)",
        type=["xlsx"],
        help="Maps Seller SKU → OMS SKU for PWN lookup only"
    )
    st.markdown("---")
    st.markdown("### ⚙️ Settings")
    c1, c2 = st.columns(2)
    fixed_fee = c1.number_input("Fixed Fee (₹)", value=5, min_value=0, step=1)
    gst_rate  = c2.number_input("GST (%)", value=18, min_value=0, step=1) / 100

    st.markdown("---")
    st.markdown("""
<details>
<summary style="cursor:pointer;font-weight:600;color:#ffd700">📖 Brand Detection Rules</summary>
<div style="font-size:0.78rem;margin-top:6px;line-height:1.7">

| Product starts with | Brand |
|---------------------|-------|
| Yash Gallery | Yash Gallery |
| KALINI | KALINI |
| Tasrika | Tasrika |
| AKIKO | Pushpa |
| HouseOfCommon | Pushpa |
| IKRASS | IKRASS |

</div>
</details>

<details style="margin-top:8px">
<summary style="cursor:pointer;font-weight:600;color:#ffd700">📖 Calculation Logic</summary>
<div style="font-size:0.78rem;margin-top:6px;line-height:1.7">
<b>GT</b> → slab on Invoice Amount<br>
<b>Selling Price</b> = Invoice − GT<br>
<b>Taxable Value</b> = SP − SP/105×5<br>
<b>TDS</b> = TV × 0.1% | <b>TCS</b> = TV × 0.5%<br>
<b>Received</b> = SP − Charges − GST − TDS − TCS<br>
<b>Difference</b> = Received − (Qty × PWN)
</div>
</details>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def extract_brand_from_product(product: str) -> str:
    if not product or str(product).strip().lower() in ("nan",""):
        return ""
    p = str(product).strip()
    for brand in KNOWN_BRANDS:
        if p.lower().startswith(brand.lower()):
            return BRAND_DISPLAY_MAP.get(brand, brand)
    return ""

def strip_vendor_prefix(sku: str) -> str:
    upper = sku.upper()
    for p in VENDOR_PREFIXES:
        if upper.startswith(p.upper()):
            sku = sku[len(p):]
            break
    sku = re.sub(r"(?i)YKN", "YK", sku)
    sku = re.sub(r"(?i)YKC", "YK", sku)
    return sku

def get_sku_base(sku: str) -> str:
    key = re.sub(r"\s*-\s*", "-", sku.strip().upper())
    parts = key.rsplit("-", 1)
    return parts[0] if len(parts) == 2 else key

def lookup_sub_cat(raw_sku: str, sku_info_dict: dict) -> tuple:
    key_raw = raw_sku.strip().upper()
    info = sku_info_dict.get(key_raw)
    if info and info.get("sub_cat") and str(info["sub_cat"]).lower() != "nan":
        return info["sub_cat"], "exact"

    stripped = strip_vendor_prefix(raw_sku).strip().upper()
    if stripped != key_raw:
        info2 = sku_info_dict.get(stripped)
        if info2 and info2.get("sub_cat") and str(info2["sub_cat"]).lower() != "nan":
            return info2["sub_cat"], "exact-stripped"

    base_raw = get_sku_base(key_raw)
    for csk, ci in sku_info_dict.items():
        if get_sku_base(csk) == base_raw and ci.get("sub_cat") and str(ci["sub_cat"]).lower() != "nan":
            return ci["sub_cat"], f"base({csk})"

    if stripped != key_raw:
        base_str = get_sku_base(stripped)
        for csk, ci in sku_info_dict.items():
            if get_sku_base(csk) == base_str and ci.get("sub_cat") and str(ci["sub_cat"]).lower() != "nan":
                return ci["sub_cat"], f"base-stripped({csk})"

    return "", "not_found"

def lookup_pwn(sku: str, pwn_dict: dict) -> tuple:
    key = sku.strip().upper()
    val = pwn_dict.get(key)
    if val is not None and pd.notna(val): return float(val), "direct"
    parts = key.rsplit("-", 1)
    if len(parts) == 2:
        base, size = parts
        for ps in ORDER_TO_PRICE_SIZE.get(size, []):
            val = pwn_dict.get(f"{base}-{ps.upper()}")
            if val is not None and pd.notna(val): return float(val), f"size-expand({ps})"
    base = get_sku_base(key)
    if base:
        for csk, cv in pwn_dict.items():
            if get_sku_base(csk) == base and pd.notna(cv): return float(cv), f"base-match({csk})"
    return np.nan, "not_found"

def lookup_pwn_with_replace(sku, pwn_dict, replace_map):
    v, m = lookup_pwn(sku, pwn_dict)
    if m != "not_found": return v, m
    oms = replace_map.get(sku.strip().upper())
    if oms:
        v2, m2 = lookup_pwn(oms, pwn_dict)
        if m2 != "not_found": return v2, f"replace→{m2}"
    return np.nan, "not_found"

# ── Slab lookups ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _build_brand_cat_index(charges_df_json: str):
    """Pre-build index for O(1) brand+category slab lookup — called once per charges file."""
    import json
    df = pd.read_json(charges_df_json, orient="records")
    idx = {}
    for _, r in df.iterrows():
        bn = str(r.get("Brand Name","")).strip().lower()
        cn = str(r.get("Category","")).strip().lower()
        if bn and cn and bn != "nan" and cn != "nan":
            idx.setdefault((bn, cn), []).append(r.to_dict())
    return idx

def _filter_brand_cat(charges_df, brand, cat):
    if not brand or not cat: return pd.DataFrame()
    if "Brand Name" not in charges_df.columns or "Category" not in charges_df.columns:
        return pd.DataFrame()
    bn = str(brand).strip().lower()
    cn = str(cat).strip().lower()
    if bn == "nan" or cn == "nan": return pd.DataFrame()
    bm = charges_df["Brand Name"].fillna("").astype(str).str.strip().str.lower() == bn
    cm = charges_df["Category"].fillna("").astype(str).str.strip().str.lower() == cn
    return charges_df[bm & cm].copy()

def lookup_gt(brand, cat, inv_amount, charges_df):
    for _, r in _filter_brand_cat(charges_df, brand, cat).iterrows():
        lo, hi, gt = r.get("GT Lower Limit"), r.get("GT Upper Limit"), r.get("GT Charge")
        if pd.isna(lo) or pd.isna(hi) or pd.isna(gt): continue
        try:
            if float(lo) <= inv_amount <= float(hi) + 0.99: return float(gt)
        except: continue
    return np.nan

def lookup_commission(brand, cat, sell_price, charges_df):
    for _, r in _filter_brand_cat(charges_df, brand, cat).iterrows():
        lo, hi, ch = r.get("Lower Limit Commision"), r.get("Upper Limit Commision"), r.get("Commision Charge")
        if pd.isna(lo) or pd.isna(hi) or pd.isna(ch): continue
        try:
            if float(lo) <= sell_price <= float(hi) + 0.99: return round(float(ch) * sell_price, 5)
        except: continue
    return np.nan

def lookup_collection(brand, cat, sell_price, charges_df):
    for _, r in _filter_brand_cat(charges_df, brand, cat).iterrows():
        lo_raw, hi, cf = r.get("Collection Lower Limit"), r.get("Collection Upper Limit"), r.get("Collection Charge")
        if pd.isna(hi) or pd.isna(cf): continue
        try:
            cf_val = float(cf) if pd.notna(cf) else 0.0
            lo_val = 0.0 if (pd.isna(lo_raw) or str(lo_raw).strip().startswith(">")) else float(lo_raw)
            if lo_val < sell_price <= float(hi) + 0.99: return round(cf_val * sell_price, 5)
        except: continue
    return np.nan

# ══════════════════════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def parse_charges_df_cached(file_bytes: bytes) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=None)
    return _parse_charges_df(raw)

def _parse_charges_df(raw):
    df = raw.copy()
    df.columns = [str(c).strip() for c in raw.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)
    if "Brand Name" in df.columns: df["Brand Name"] = df["Brand Name"].ffill()
    if "Category"   in df.columns: df["Category"]   = df["Category"].ffill()
    df = df[df["Category"].notna()].copy()
    for col in ["Lower Limit Commision","Upper Limit Commision","Commision Charge",
                "Collection Lower Limit","Collection Upper Limit",
                "GT Lower Limit","GT Upper Limit","GT Charge"]:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce")
    if "Collection Charge" in df.columns:
        df["Collection Charge"] = (df["Collection Charge"].astype(str)
            .str.replace("₹","",regex=False).str.strip()
            .pipe(pd.to_numeric, errors="coerce"))
    return df

@st.cache_data(show_spinner=False)
def parse_sku_info_cached(file_bytes: bytes, sheet_idx: int) -> dict:
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_idx, header=None)
    return _parse_sku_info(raw)

def _parse_sku_info(raw):
    df = raw.copy()
    df.columns = [str(c).strip() for c in raw.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)
    sku_info = {}
    for _, row in df.iterrows():
        sku     = str(row.get("Seller SKU Id","")).strip().upper()
        sub_cat = str(row.get("Sub-category","")).strip()
        brand   = str(row.get("Brand","")).strip() if "Brand" in df.columns else ""
        if sku:
            sku_info[sku] = {"sub_cat": sub_cat, "brand": brand}
    return sku_info

@st.cache_data(show_spinner=False)
def parse_pwn_dict_cached(file_bytes: bytes, sheet_idx: int) -> dict:
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_idx, header=None)
    return _parse_pwn_dict(raw)

def _parse_pwn_dict(raw):
    df = raw.copy()
    df.columns = [str(c).strip() for c in raw.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)
    df["OMS Child SKU"] = df["OMS Child SKU"].astype(str).str.strip()
    df["PWN+10%+50"] = pd.to_numeric(df["PWN+10%+50"], errors="coerce")
    return dict(zip(df["OMS Child SKU"].str.upper(), df["PWN+10%+50"]))

@st.cache_data(show_spinner=False)
def parse_replace_map_cached(file_bytes: bytes) -> dict:
    xl = pd.read_excel(BytesIO(file_bytes), header=None)
    df = xl.copy()
    df.columns = [str(c).strip() for c in xl.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)
    return dict(zip(
        df["Seller SKU Id"].astype(str).str.strip().str.upper(),
        df["OMS SKU"].astype(str).str.strip().str.upper()
    ))

# ══════════════════════════════════════════════════════════════════════════════
# FILE READER
# ══════════════════════════════════════════════════════════════════════════════
def read_order_file(f):
    name = f.name.lower()
    try:
        if name.endswith(".csv"):
            raw = f.read()
            df = None
            for enc in ("utf-8","utf-8-sig","latin-1","cp1252"):
                try:
                    df = pd.read_csv(BytesIO(raw), encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                return pd.DataFrame(), f"Could not decode '{f.name}'."
        elif name.endswith((".xlsx",".xls")):
            engine = "openpyxl" if name.endswith(".xlsx") else "xlrd"
            df = pd.read_excel(f, engine=engine)
        else:
            return pd.DataFrame(), f"Unsupported: '{f.name}'"

        df.columns = [str(c).strip() for c in df.columns]
        missing = REQUIRED_ORDER_COLS - set(df.columns)
        if missing:
            return pd.DataFrame(), f"'{f.name}' missing columns: {', '.join(sorted(missing))}"
        df["_source_file"] = f.name
        return df, ""
    except Exception as e:
        return pd.DataFrame(), f"Error reading '{f.name}': {e}"

def load_all_order_files(files):
    frames, file_info, errors = [], [], []
    for f in files:
        df, err = read_order_file(f)
        if err:
            errors.append(err)
            file_info.append({"File": f.name, "Rows": 0, "Status": "❌ Error"})
        else:
            frames.append(df)
            file_info.append({"File": f.name, "Rows": len(df), "Status": "✅ OK"})
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return combined, file_info, errors

# ══════════════════════════════════════════════════════════════════════════════
# CORE RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════
def run_reconciliation(order_df, charges_df, sku_info_dict, pwn_dict,
                       fixed_fee, gst_rate,
                       replace_map=None, pwn_overrides=None, sku_corrections=None):
    replace_map     = replace_map     or {}
    pwn_overrides   = pwn_overrides   or {}
    sku_corrections = sku_corrections or {}
    rows_out = []

    for _, row in order_df.iterrows():
        raw_sku    = str(row.get("SKU", "")).strip()
        product    = str(row.get("Product", "")).strip()
        order_id   = str(row.get("Order Id", "")).strip()
        ordered_on = row.get("Ordered On", "")
        inv_amount = float(row.get("Invoice Amount", 0) or 0)
        quantity   = int(row.get("Quantity", 1) or 1)
        source_file = str(row.get("_source_file", ""))

        # Step 1: Manual SKU correction
        corrected_raw = sku_corrections.get(raw_sku.upper(), raw_sku)

        # Step 2: Replace SKU mapping (for general lookups)
        if corrected_raw.strip().upper() in replace_map:
            corrected_raw = replace_map[corrected_raw.strip().upper()]

        # Step 3: Brand from Product name
        brand_name = extract_brand_from_product(product)

        # Step 4: Sub-category from Sheet 1
        sub_cat, cat_match_note = lookup_sub_cat(corrected_raw, sku_info_dict)
        cat = sub_cat.strip() if sub_cat and str(sub_cat).lower() != "nan" else ""

        # Step 5: Stripped SKU
        sku_for_pwn = strip_vendor_prefix(corrected_raw)

        # Step 5b: Separate SKU for PWN only
        sku_for_pwn_replace = sku_for_pwn
        _stripped_upper = sku_for_pwn.strip().upper()
        if _stripped_upper in replace_map:
            sku_for_pwn_replace = replace_map[_stripped_upper]
        else:
            _corrected_upper = corrected_raw.strip().upper()
            if _corrected_upper in replace_map:
                sku_for_pwn_replace = replace_map[_corrected_upper]
            else:
                _raw_upper = raw_sku.strip().upper()
                if _raw_upper in replace_map:
                    sku_for_pwn_replace = replace_map[_raw_upper]

        # Step 6: Slab lookups
        gt_val = sell_price = commission = coll_fee = np.nan
        charge_method = "not_found"

        if brand_name and cat:
            gt_val = lookup_gt(brand_name, cat, inv_amount, charges_df)
            if pd.notna(gt_val):
                sell_price = round(inv_amount - gt_val, 5)
                commission = lookup_commission(brand_name, cat, sell_price, charges_df)
                coll_fee   = lookup_collection(brand_name, cat, sell_price, charges_df)
                if pd.notna(commission) and pd.notna(coll_fee):
                    charge_method = f"{brand_name} | {cat}"
                else:
                    gt_val = sell_price = commission = coll_fee = np.nan

        # Step 7: Final amounts
        if pd.isna(gt_val):
            sell_price = gt_val = commission = coll_fee = np.nan
            total_charges = gst_on_charges = taxable_value = np.nan
            tds = tcs = total_deductions = received_amount = np.nan
            if not brand_name:
                charge_method = "no_brand_in_product"
            elif not cat:
                charge_method = f"no_subcat|brand={brand_name}"
            else:
                charge_method = f"slab_missing|{brand_name}|{cat}"
        else:
            commission    = commission if pd.notna(commission) else 0.0
            coll_fee      = coll_fee   if pd.notna(coll_fee)   else 0.0
            total_charges  = round(commission + coll_fee + float(fixed_fee), 5)
            gst_on_charges = round(total_charges * gst_rate, 5)
            taxable_value  = round(sell_price - (sell_price / 105 * 5), 5)
            tds            = round(taxable_value * TDS_RATE, 5)
            tcs            = round(taxable_value * TCS_RATE, 5)
            total_deductions = round(total_charges + gst_on_charges + tds + tcs, 5)
            received_amount  = round(sell_price - total_charges - gst_on_charges - tds - tcs, 5)

        # Step 8: PWN lookup
        pwn_val, match_method = lookup_pwn(sku_for_pwn_replace, pwn_dict)
        if match_method == "not_found" and sku_for_pwn_replace != sku_for_pwn:
            pwn_val, match_method = lookup_pwn(sku_for_pwn, pwn_dict)
        if match_method == "not_found":
            pwn_val, match_method = lookup_pwn_with_replace(sku_for_pwn, pwn_dict, replace_map)

        # Manual PWN override (highest priority)
        if sku_for_pwn_replace.upper() in pwn_overrides:
            pwn_val, match_method = float(pwn_overrides[sku_for_pwn_replace.upper()]), "manual"
        elif sku_for_pwn.upper() in pwn_overrides:
            pwn_val, match_method = float(pwn_overrides[sku_for_pwn.upper()]), "manual"

        full_match_note = match_method
        if cat_match_note and cat_match_note not in ("exact","exact-stripped"):
            full_match_note = f"{match_method} | cat:{cat_match_note}"

        # Step 9: Difference
        if pd.notna(received_amount) and pd.notna(pwn_val):
            pwn_benchmark = round(pwn_val * quantity, 5)
            difference    = round(received_amount - pwn_benchmark, 5)
        else:
            pwn_benchmark = difference = np.nan

        rows_out.append({
            "Order Id":         order_id,
            "SKU":              raw_sku,
            "SKU for PWN":      sku_for_pwn_replace,
            "Product":          product,
            "Brand Name":       brand_name,
            "Ordered On":       ordered_on,
            "Sub-Category":     sub_cat,
            "Charge Method":    charge_method,
            "Qty":              quantity,
            "Invoice Amount":   inv_amount,
            "GT (As Per Calc)": gt_val,
            "Selling Price":    sell_price,
            "Commission":       commission,
            "Collection Fee":   coll_fee,
            "Fixed Fee":        float(fixed_fee),
            "Total Charges":    total_charges,
            "GST on Charges":   gst_on_charges,
            "Taxable Value":    taxable_value,
            "TDS":              tds,
            "TCS":              tcs,
            "Total Deductions": total_deductions,
            "Received Amount":  received_amount,
            "PWN":              pwn_val,
            "PWN Benchmark":    pwn_benchmark,
            "PWN Match":        full_match_note,
            "Difference":       difference,
            "Source File":      source_file,
        })

    return pd.DataFrame(rows_out)

# ══════════════════════════════════════════════════════════════════════════════
# FORMATTING
# ══════════════════════════════════════════════════════════════════════════════
def fmt_inr(x):
    try:
        if pd.isna(x): return "—"
        return f"₹{float(x):,.2f}"
    except: return str(x)

def style_table(df, diff_col="Difference"):
    fmt_dict = {c: fmt_inr for c in df.columns if c in MONEY_COLS}
    def colour_diff(val):
        try:
            v = float(val)
            if v < 0:  return "color: #c0392b; font-weight: 700"
            if v > 0:  return "color: #1e8449; font-weight: 700"
            return "color: #7d6608; font-weight: 600"
        except: return ""
    styler = df.style.format(fmt_dict)
    if diff_col in df.columns:
        styler = styler.map(colour_diff, subset=[diff_col])
    return styler

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def apply_roc_sheet_style(ws, df):
    C_HEADER_BG = "1A3C5E"; C_HEADER_FG = "FFFFFF"
    C_ALT1 = "EAF2FB"; C_ALT2 = "FFFFFF"
    C_GREEN_BG = "D6EFDD"; C_RED_BG = "FDDEDE"; C_ZERO_BG = "FFF9E6"
    C_TOTAL_BG = "1A3C5E"; C_TOTAL_FG = "FFD700"; C_BORDER = "B0C4D8"
    thin  = Side(style="thin",   color=C_BORDER)
    thick = Side(style="medium", color="1A3C5E")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_header = Border(left=thick, right=thick, top=thick, bottom=thick)
    money_names = set(MONEY_COLS)
    cols = df.columns.tolist()
    C = {name: get_column_letter(i+1) for i, name in enumerate(cols)}
    col_widths = {
        "Order Id":20,"SKU":28,"SKU for PWN":28,"Product":40,"Brand Name":18,
        "Ordered On":14,"Sub-Category":20,"Charge Method":28,"Qty":6,
        "Invoice Amount":15,"GT (As Per Calc)":15,"Selling Price":15,
        "Commission":14,"Collection Fee":15,"Fixed Fee":10,"Total Charges":15,
        "GST on Charges":15,"Taxable Value":14,"TDS":10,"TCS":10,
        "Total Deductions":16,"Received Amount":16,"PWN":12,
        "PWN Benchmark":15,"PWN Match":16,"Difference":14,"Source File":20,
    }
    for i, cn in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(cn, 14)
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.font      = Font(bold=True, color=C_HEADER_FG, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr_header
    ws.row_dimensions[1].height = 30
    diff_col_idx = cols.index("Difference")+1 if "Difference" in cols else None

    def row_formulas(r):
        sp=C.get("Selling Price",""); inv=C.get("Invoice Amount","")
        gt=C.get("GT (As Per Calc)",""); qty=C.get("Qty","")
        com=C.get("Commission",""); colf=C.get("Collection Fee","")
        ff=C.get("Fixed Fee",""); tc=C.get("Total Charges","")
        gst=C.get("GST on Charges",""); tv=C.get("Taxable Value","")
        tds=C.get("TDS",""); tcs=C.get("TCS","")
        td=C.get("Total Deductions",""); ra=C.get("Received Amount","")
        pwn=C.get("PWN",""); pb=C.get("PWN Benchmark","")
        diff=C.get("Difference","")
        fmls = {}
        if sp and inv and gt:   fmls["Selling Price"]=f'=IF(OR({gt}{r}="",{gt}{r}=0),"",ROUND({inv}{r}-{gt}{r},2))'
        if tc and com and colf: fmls["Total Charges"]=f'=IF({sp}{r}="","",ROUND({com}{r}+{colf}{r}+{ff}{r},2))'
        if gst and tc:          fmls["GST on Charges"]=f'=IF({tc}{r}="","",ROUND({tc}{r}*0.18,2))'
        if tv and sp:           fmls["Taxable Value"]=f'=IF({sp}{r}="","",ROUND({sp}{r}-{sp}{r}/105*5,2))'
        if tds and tv:          fmls["TDS"]=f'=IF({tv}{r}="","",ROUND({tv}{r}*0.001,2))'
        if tcs and tv:          fmls["TCS"]=f'=IF({tv}{r}="","",ROUND({tv}{r}*0.005,2))'
        if td and tc:           fmls["Total Deductions"]=f'=IF({tc}{r}="","",ROUND({tc}{r}+{gst}{r}+{tds}{r}+{tcs}{r},2))'
        if ra and sp and tc:    fmls["Received Amount"]=f'=IF({sp}{r}="","",ROUND({sp}{r}-{tc}{r}-{gst}{r}-{tds}{r}-{tcs}{r},2))'
        if pb and pwn and qty:  fmls["PWN Benchmark"]=f'=IF({pwn}{r}="","",ROUND({pwn}{r}*{qty}{r},2))'
        if diff and ra and pb:  fmls["Difference"]=f'=IF(OR({ra}{r}="",{pb}{r}=""),"",ROUND({ra}{r}-{pb}{r},2))'
        return fmls

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        alt_fill = PatternFill("solid", fgColor=C_ALT1 if r_idx % 2 == 0 else C_ALT2)
        fmls = row_formulas(r_idx)
        for c_idx, (col_name, val) in enumerate(zip(cols, row_data), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if col_name in fmls:
                cell.value = fmls[col_name]
            else:
                cell.value = None if (isinstance(val, float) and np.isnan(val)) else val
            cell.border = bdr
            cell.font   = Font(size=9, name="Calibri")
            cell.fill   = alt_fill
            if col_name in money_names:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Qty":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if c_idx == diff_col_idx:
                try:
                    v = float(val)
                    if not np.isnan(v):
                        if v < 0:   cell.fill = PatternFill("solid", fgColor=C_RED_BG);   cell.font = Font(color="C0392B", bold=True, size=9, name="Calibri")
                        elif v > 0: cell.fill = PatternFill("solid", fgColor=C_GREEN_BG); cell.font = Font(color="1E8449", bold=True, size=9, name="Calibri")
                        else:       cell.fill = PatternFill("solid", fgColor=C_ZERO_BG);  cell.font = Font(color="7D6608", bold=True, size=9, name="Calibri")
                except: pass
        ws.row_dimensions[r_idx].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    last_data_row = len(df) + 1
    total_row = last_data_row + 2
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.fill   = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell.font   = Font(bold=True, color=C_TOTAL_FG, size=10, name="Calibri")
        cell.border = bdr_header
        if c_idx == 1:
            cell.value = "TOTALS"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_name in money_names:
            col_l = get_column_letter(c_idx)
            cell.value = f"=SUM({col_l}2:{col_l}{last_data_row})"
            cell.number_format = '₹#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[total_row].height = 22

def apply_summary_style(ws):
    thin = Side(style="thin", color="AED6F1")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="2C3E50")
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[1].height = 24
    for r_idx in range(2, ws.max_row + 1):
        fill = PatternFill("solid", fgColor="EBF5FB" if r_idx % 2 == 0 else "FFFFFF")
        for cell in ws[r_idx]:
            cell.fill = fill; cell.font = Font(size=9, name="Calibri")
            cell.border = bdr; cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r_idx].height = 15
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 40)
    last_row = ws.max_row; total_row = last_row + 2
    for c_idx in range(1, ws.max_column + 1):
        sc = ws.cell(row=2, column=c_idx)
        tc = ws.cell(row=total_row, column=c_idx)
        tc.fill   = PatternFill("solid", fgColor="2C3E50")
        tc.font   = Font(bold=True, color="FFD700", size=10, name="Calibri")
        tc.border = Border(
            left=Side(style="medium", color="2C3E50"), right=Side(style="medium", color="2C3E50"),
            top=Side(style="medium",  color="2C3E50"), bottom=Side(style="medium", color="2C3E50")
        )
        if c_idx == 1:
            tc.value = "TOTALS"
        elif isinstance(sc.value, (int, float)):
            cl = get_column_letter(c_idx)
            tc.value = f"=SUM({cl}2:{cl}{last_row})"
            tc.number_format = '₹#,##0.00'
            tc.alignment = Alignment(horizontal="right", vertical="center")
    ws.freeze_panes = "A2"

def to_excel(recon_df, summary_df, cat_df, brand_df=None):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        recon_df.to_excel(writer, index=False, sheet_name="Reconciliation")
        cat_df.to_excel(writer, index=False, sheet_name="Category Breakdown")
        summary_df.to_excel(writer, index=False, sheet_name="Charges Summary")
        if brand_df is not None and not brand_df.empty:
            brand_df.to_excel(writer, index=False, sheet_name="Brand Breakdown")
        apply_roc_sheet_style(writer.sheets["Reconciliation"], recon_df)
        apply_summary_style(writer.sheets["Category Breakdown"])
        apply_summary_style(writer.sheets["Charges Summary"])
        if brand_df is not None and not brand_df.empty:
            apply_summary_style(writer.sheets["Brand Breakdown"])
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
def build_summary(df):
    valid = df[df["Received Amount"].notna()]
    totals = {"Metric": [], "Value": []}
    for label, val in [
        ("Total Orders",            len(df)),
        ("Orders Calculated",       int(df["Received Amount"].notna().sum())),
        ("Orders Skipped (no match)", int(df["Received Amount"].isna().sum())),
        ("Total Invoice Amount",    df["Invoice Amount"].sum()),
        ("Total GT (As Per Calc)", valid["GT (As Per Calc)"].sum()),
        ("Total Selling Price",    valid["Selling Price"].sum()),
        ("Total Commission",       valid["Commission"].sum()),
        ("Total Collection Fee",   valid["Collection Fee"].sum()),
        ("Total Fixed Fee",        valid["Fixed Fee"].sum()),
        ("Total Charges",          valid["Total Charges"].sum()),
        ("Total GST on Charges",   valid["GST on Charges"].sum()),
        ("Total TDS",              valid["TDS"].sum()),
        ("Total TCS",              valid["TCS"].sum()),
        ("Total Deductions",       valid["Total Deductions"].sum()),
        ("Total Received Amount",  valid["Received Amount"].sum()),
        ("Net Difference vs PWN",  valid["Difference"].sum()),
        ("Orders Negative Diff",   int((valid["Difference"] < 0).sum())),
        ("Orders Positive Diff",   int((valid["Difference"] > 0).sum())),
        ("Orders – No PWN found",  int(df["Difference"].isna().sum())),
        ("Avg Received per Order", valid["Received Amount"].mean()),
        ("Avg Difference per Order", valid["Difference"].mean()),
    ]:
        totals["Metric"].append(label)
        totals["Value"].append(round(val, 2) if isinstance(val, float) else val)

    summary_df = pd.DataFrame(totals)

    cat_df = (
        valid.groupby("Sub-Category")
        .agg(
            Orders=("Order Id","count"),
            Invoice_Total=("Invoice Amount","sum"),
            GT_Total=("GT (As Per Calc)","sum"),
            Selling_Total=("Selling Price","sum"),
            Commission=("Commission","sum"),
            Collection=("Collection Fee","sum"),
            Fixed=("Fixed Fee","sum"),
            Total_Charges=("Total Charges","sum"),
            GST_Total=("GST on Charges","sum"),
            TDS_Total=("TDS","sum"),
            TCS_Total=("TCS","sum"),
            Deductions=("Total Deductions","sum"),
            Received_Total=("Received Amount","sum"),
            Net_Diff=("Difference","sum"),
            Avg_Diff=("Difference","mean"),
        )
        .reset_index()
        .sort_values("Invoice_Total", ascending=False)
        .round(2)
    )
    cat_df.columns = [
        "Sub-Category","Orders","Invoice Total","GT Total","Selling Total",
        "Commission","Collection Fee","Fixed Fee","Total Charges",
        "GST Total","TDS Total","TCS Total","Total Deductions",
        "Received Total","Net Difference","Avg Difference",
    ]

    # NEW: Brand-level breakdown
    brand_df = (
        valid.groupby("Brand Name")
        .agg(
            Orders=("Order Id","count"),
            Invoice_Total=("Invoice Amount","sum"),
            Selling_Total=("Selling Price","sum"),
            Total_Charges=("Total Charges","sum"),
            Deductions=("Total Deductions","sum"),
            Received_Total=("Received Amount","sum"),
            Net_Diff=("Difference","sum"),
        )
        .reset_index()
        .sort_values("Invoice_Total", ascending=False)
        .round(2)
    )
    brand_df.columns = [
        "Brand","Orders","Invoice Total","Selling Total",
        "Total Charges","Total Deductions","Received Total","Net Difference",
    ]

    return summary_df, cat_df, brand_df

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for k, v in [
    ("pwn_overrides", {}), ("sku_corrections", {}), ("result_df", None),
    ("charges_df", None), ("sku_info_dict", {}), ("pwn_dict", {}),
    ("order_df", None), ("replace_map", {}),
]:
    if k not in st.session_state: st.session_state[k] = v

# ══════════════════════════════════════════════════════════════════════════════
# DISPLAY COLUMNS
# ══════════════════════════════════════════════════════════════════════════════
DISPLAY_COLS = [
    "Order Id","SKU","SKU for PWN","Product","Brand Name","Ordered On",
    "Sub-Category","Charge Method","Qty","Invoice Amount","GT (As Per Calc)",
    "Selling Price","Commission","Collection Fee","Fixed Fee","Total Charges",
    "GST on Charges","Taxable Value","TDS","TCS","Total Deductions",
    "Received Amount","PWN","PWN Benchmark","Difference","PWN Match","Source File",
]

# ══════════════════════════════════════════════════════════════════════════════
# MAIN LOGIC
# ══════════════════════════════════════════════════════════════════════════════
if order_files and charges_file:

    # ── Load & parse files ────────────────────────────────────────────────────
    with st.spinner("⏳ Reading and parsing files…"):
        order_df, file_info, file_errors = load_all_order_files(order_files)

        # Sidebar file summary
        st.sidebar.markdown("---")
        st.sidebar.markdown("**📄 Uploaded Order Files**")
        st.sidebar.dataframe(pd.DataFrame(file_info), hide_index=True, use_container_width=True)
        st.sidebar.caption(f"Total rows loaded: **{len(order_df):,}**")
        for err in file_errors:
            st.warning(f"⚠️ {err}")
        if order_df.empty:
            st.error("❌ No valid order rows loaded.")
            st.stop()

        # Read charges file once, use cached parsers
        charges_bytes = charges_file.read()
        xl_raw = pd.read_excel(BytesIO(charges_bytes), sheet_name=None, header=None)
        sheets = list(xl_raw.values())
        if len(sheets) < 3:
            st.error(f"❌ Data Excel needs ≥3 sheets. Found: {list(xl_raw.keys())}")
            st.stop()

        charges_df    = _parse_charges_df(sheets[0])
        sku_info_dict = _parse_sku_info(sheets[1])
        pwn_dict      = _parse_pwn_dict(sheets[2])
        replace_map   = parse_replace_map_cached(replace_sku_file.read()) if replace_sku_file else {}

        if replace_sku_file:
            st.sidebar.success(f"✅ Replace SKU loaded: {len(replace_map):,} entries")

        known_brands = (
            sorted(charges_df["Brand Name"].dropna().unique().tolist())
            if "Brand Name" in charges_df.columns else []
        )
        st.sidebar.markdown(f"**Brands in Sheet 0:** `{'`, `'.join(known_brands)}`")

        st.session_state.update({
            "charges_df": charges_df, "sku_info_dict": sku_info_dict,
            "pwn_dict": pwn_dict, "order_df": order_df, "replace_map": replace_map,
        })

    # ── Debug expanders ───────────────────────────────────────────────────────
    with st.expander("🔍 Brand & Sub-Category Detection — first 20 rows", expanded=False):
        sample = order_df[["SKU","Product"]].head(20).copy()
        sample["Brand (from Product)"] = sample["Product"].apply(extract_brand_from_product)
        sample["Sub-Category"] = sample["SKU"].apply(lambda s: lookup_sub_cat(s, sku_info_dict)[0])
        sample["Charge Rows"] = sample.apply(
            lambda r: len(_filter_brand_cat(charges_df, r["Brand (from Product)"], r["Sub-Category"])), axis=1
        )
        sample["Ready"] = sample["Charge Rows"].apply(lambda x: "✅" if x > 0 else "❌")
        st.dataframe(sample, use_container_width=True)
        cats = sorted(charges_df["Category"].dropna().unique().tolist()) if "Category" in charges_df.columns else []
        col_a, col_b = st.columns(2)
        col_a.markdown(f"**Brands in Sheet 0:** {', '.join(known_brands)}")
        col_b.markdown(f"**Categories in Sheet 0:** {', '.join(cats)}")

    with st.expander("📊 Sheet Structure Preview", expanded=False):
        t0, t1, t2 = st.tabs(["Sheet 0 · Charges", "Sheet 1 · SKU→Category", "Sheet 2 · PWN Prices"])
        with t0: st.dataframe(charges_df.head(15), use_container_width=True)
        with t1:
            s1 = pd.DataFrame([{"SKU": k, "Sub-Cat": v.get("sub_cat","")} for k, v in list(sku_info_dict.items())[:15]])
            st.dataframe(s1, use_container_width=True)
        with t2:
            s2 = pd.DataFrame([{"SKU": k, "PWN": v} for k, v in list(pwn_dict.items())[:15]])
            st.dataframe(s2, use_container_width=True)

    # ── Run reconciliation ────────────────────────────────────────────────────
    with st.spinner("🔄 Running reconciliation…"):
        result_df = run_reconciliation(
            st.session_state["order_df"],
            st.session_state["charges_df"],
            st.session_state["sku_info_dict"],
            st.session_state["pwn_dict"],
            fixed_fee, gst_rate,
            replace_map=st.session_state["replace_map"],
            pwn_overrides=st.session_state["pwn_overrides"],
            sku_corrections=st.session_state["sku_corrections"],
        )
    st.session_state["result_df"] = result_df
    summary_df, cat_df, brand_df = build_summary(result_df)

    calc_count   = int(result_df["Received Amount"].notna().sum())
    skip_count   = int(result_df["Received Amount"].isna().sum())
    replace_resolved = result_df[result_df["PWN Match"].str.startswith("replace", na=False)]

    st.success(
        f"✅ **{len(result_df):,}** orders processed  |  "
        f"**{calc_count:,}** calculated  |  "
        f"**{skip_count:,}** skipped"
        + (f"  |  **{len(replace_resolved):,}** PWN via Replace SKU" if len(replace_resolved) else "")
    )

    # ══════════════════════════════════════════════════════════════════════════
    # TABS
    # ══════════════════════════════════════════════════════════════════════════
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋  Reconciliation",
        "💰  Charges Summary",
        "📊  Category Breakdown",
        "🏷️  Brand Breakdown",
    ])

    # ── TAB 1: RECONCILIATION ─────────────────────────────────────────────────
    with tab1:

        # SKU issues panel
        broken_df = result_df[
            result_df["Received Amount"].isna() | (result_df["PWN Match"] == "not_found")
        ]
        if len(broken_df):
            no_cat = int(broken_df["Received Amount"].isna().sum())
            no_pwn = int((broken_df["PWN Match"] == "not_found").sum())
            with st.expander(
                f"✏️  **{len(broken_df)} SKU(s) have lookup issues** — click to fix",
                expanded=False
            ):
                st.info("Enter the corrected SKU to re-run all lookups for that row.")
                st.caption(
                    f"🔴 No category/GT: **{no_cat}**  |  "
                    f"🟡 No PWN: **{no_pwn}**"
                )
                broken_skus = broken_df["SKU"].unique().tolist()
                correction_inputs = {}
                h1, h2, h3, h4 = st.columns([3, 2, 2, 3])
                h1.markdown("**Original SKU**"); h2.markdown("**Issue**")
                h3.markdown("**Corrected SKU**"); h4.markdown("**Live preview**")
                st.markdown("---")
                for sku in broken_skus:
                    sku_rows = broken_df[broken_df["SKU"] == sku]
                    issues = []
                    if sku_rows["Received Amount"].isna().any(): issues.append("❌ No category/GT")
                    if (sku_rows["PWN Match"] == "not_found").any(): issues.append("⚠️ No PWN")
                    existing = st.session_state["sku_corrections"].get(sku.upper(), "")
                    c1, c2, c3, c4 = st.columns([3, 2, 2, 3])
                    c1.markdown(
                        f"<div style='padding-top:6px;font-size:0.88rem;word-break:break-all'>"
                        f"<code>{sku}</code></div>", unsafe_allow_html=True
                    )
                    c2.markdown(
                        f"<div style='padding-top:6px;font-size:0.82rem'>"
                        f"{'  &  '.join(issues)}</div>", unsafe_allow_html=True
                    )
                    corrected = c3.text_input(
                        "Corrected SKU", value=existing,
                        placeholder="e.g. YK1234-L",
                        label_visibility="collapsed", key=f"sku_corr_{sku}"
                    )
                    correction_inputs[sku] = corrected.strip()
                    if existing:
                        sc_p, _ = lookup_sub_cat(existing, st.session_state["sku_info_dict"])
                        pwn_v, _ = lookup_pwn_with_replace(
                            strip_vendor_prefix(existing),
                            st.session_state["pwn_dict"],
                            st.session_state["replace_map"]
                        )
                        parts = []
                        if sc_p and sc_p != "nan": parts.append(f"📦 Sub-cat: *{sc_p}*")
                        if pd.notna(pwn_v): parts.append(f"💰 PWN: ₹{pwn_v:,.2f}")
                        html = (
                            "<div style='padding-top:4px;font-size:0.80rem;color:#1a7a3c;line-height:1.6'>"
                            + "<br>".join(parts) + "</div>"
                        ) if parts else (
                            "<div style='padding-top:6px;font-size:0.80rem;color:#c0392b'>"
                            "⚠️ Still unresolved</div>"
                        )
                        c4.markdown(html, unsafe_allow_html=True)
                    else:
                        c4.markdown(
                            "<div style='padding-top:6px;font-size:0.80rem;color:#aaa'>"
                            "— type a correction to preview —</div>",
                            unsafe_allow_html=True
                        )
                st.markdown("---")
                cs, cc = st.columns([2, 1])
                if cs.button("💾  Save SKU Corrections & Recalculate", type="primary"):
                    st.session_state["sku_corrections"] = {
                        o.upper(): c for o, c in correction_inputs.items() if c
                    }
                    st.rerun()
                if cc.button("🗑️  Clear All Corrections"):
                    st.session_state["sku_corrections"] = {}
                    st.rerun()

        if st.session_state["sku_corrections"]:
            with st.expander(
                f"✅  **{len(st.session_state['sku_corrections'])} correction(s) active**",
                expanded=False
            ):
                st.dataframe(
                    pd.DataFrame([
                        {"Original SKU": o, "→ Corrected SKU": c}
                        for o, c in st.session_state["sku_corrections"].items()
                    ]),
                    use_container_width=True, hide_index=True
                )
                if st.button("🗑️  Clear All Active Corrections", key="clear_corr_summary"):
                    st.session_state["sku_corrections"] = {}
                    st.rerun()

        # ── KPI cards ──────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">📊 Summary</div>', unsafe_allow_html=True)
        valid = result_df[result_df["Received Amount"].notna()]
        net = valid["Difference"].sum()

        r1 = st.columns(5)
        r1[0].metric("Total Orders",     f"{len(result_df):,}")
        r1[1].metric("Invoice Total",    f"₹{result_df['Invoice Amount'].sum():,.0f}")
        r1[2].metric("GT Total",         f"₹{valid['GT (As Per Calc)'].sum():,.0f}")
        r1[3].metric("Selling Total",    f"₹{valid['Selling Price'].sum():,.0f}")
        r1[4].metric("Calculated",       f"{calc_count:,}  /  {len(result_df):,}")

        r2 = st.columns(5)
        r2[0].metric("Total Charges",    f"₹{valid['Total Charges'].sum():,.0f}")
        r2[1].metric("GST on Charges",   f"₹{valid['GST on Charges'].sum():,.0f}")
        r2[2].metric("Total TDS",        f"₹{valid['TDS'].sum():,.2f}")
        r2[3].metric("Total TCS",        f"₹{valid['TCS'].sum():,.2f}")
        r2[4].metric("Received Total",   f"₹{valid['Received Amount'].sum():,.0f}")

        # Net difference (full width metric with delta)
        nd_col, _, _ = st.columns([1, 2, 2])
        nd_col.metric(
            "Net Difference vs PWN",
            f"₹{net:,.2f}",
            delta=f"{'▲' if net >= 0 else '▼'} {abs(net):,.2f}",
            delta_color="normal" if net >= 0 else "inverse",
        )
        st.markdown("---")

        # ── Filters ───────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">🔎 Filters</div>', unsafe_allow_html=True)
        f1, f2, f3, f4, f5 = st.columns([2, 2, 2, 2, 3])

        # Multi-file filter (new!)
        source_files = sorted(result_df["Source File"].dropna().unique().tolist())
        sel_file = f5.selectbox("Source File", ["All"] + source_files) if len(source_files) > 1 else "All"

        sel_cat   = f1.selectbox("Sub-Category", ["All"] + sorted(result_df["Sub-Category"].dropna().unique().tolist()))
        sel_brand = f2.selectbox("Brand",         ["All"] + sorted(result_df["Brand Name"].dropna().unique().tolist()))
        diff_opt  = f3.selectbox("Difference type", [
            "All","Positive (+)","Negative (−)","Zero / Matched","No PWN data","No Category (NaN)"
        ])
        search = f4.text_input("🔎 SKU / Order ID")

        view = result_df.copy()
        if sel_file  != "All": view = view[view["Source File"] == sel_file]
        if sel_cat   != "All": view = view[view["Sub-Category"] == sel_cat]
        if sel_brand != "All": view = view[view["Brand Name"] == sel_brand]
        if diff_opt == "Positive (+)":      view = view[view["Difference"] > 0]
        elif diff_opt == "Negative (−)":    view = view[view["Difference"] < 0]
        elif diff_opt == "Zero / Matched":  view = view[view["Difference"] == 0]
        elif diff_opt == "No PWN data":     view = view[view["PWN Match"] == "not_found"]
        elif diff_opt == "No Category (NaN)": view = view[view["Received Amount"].isna()]
        if search.strip():
            mask = (
                view["SKU"].str.contains(search.strip(), case=False, na=False) |
                view["Order Id"].str.contains(search.strip(), case=False, na=False)
            )
            view = view[mask]

        avail_cols = [c for c in DISPLAY_COLS if c in view.columns]
        st.caption(f"Showing **{len(view):,}** of **{len(result_df):,}** orders")
        st.dataframe(
            style_table(view[avail_cols], diff_col="Difference"),
            use_container_width=True, height=500,
        )

        # ── Downloads ─────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">📥 Download</div>', unsafe_allow_html=True)
        d1, d2, d3 = st.columns(3)
        avail_full = [c for c in DISPLAY_COLS if c in result_df.columns]
        avail_filt = [c for c in DISPLAY_COLS if c in view.columns]

        d1.download_button(
            "⬇  Full Reconciliation (Excel · 4 sheets, styled)",
            data=to_excel(result_df[avail_full], summary_df, cat_df, brand_df),
            file_name="flipkart_reconciliation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        d2.download_button(
            "⬇  Filtered View (Excel · styled)",
            data=to_excel(view[avail_filt].reset_index(drop=True), summary_df, cat_df, brand_df),
            file_name="flipkart_reconciliation_filtered.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        d3.download_button(
            "⬇  Issues Only (CSV — missing SKUs)",
            data=broken_df[avail_filt].to_csv(index=False).encode() if len(broken_df) else b"No issues",
            file_name="flipkart_issues.csv",
            mime="text/csv",
        )

    # ── TAB 2: CHARGES SUMMARY ─────────────────────────────────────────────────
    with tab2:
        st.markdown('<div class="section-header">💰 Total Charges Summary</div>', unsafe_allow_html=True)
        valid = result_df[result_df["Received Amount"].notna()]
        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown("#### 📤 Flipkart Deductions")
            a1, a2 = st.columns(2)
            a1.metric("Commission",     f"₹{valid['Commission'].sum():,.2f}")
            a2.metric("Collection Fee", f"₹{valid['Collection Fee'].sum():,.2f}")
            a1.metric("Fixed Fee",      f"₹{valid['Fixed Fee'].sum():,.2f}")
            a2.metric("GST on Charges", f"₹{valid['GST on Charges'].sum():,.2f}")
            a1.metric("TDS (0.1%)",     f"₹{valid['TDS'].sum():,.2f}")
            a2.metric("TCS (0.5%)",     f"₹{valid['TCS'].sum():,.2f}")
            st.metric("🔴 Total Deductions", f"₹{valid['Total Deductions'].sum():,.2f}")

        with col_b:
            st.markdown("#### 📥 What You Receive")
            b1, b2 = st.columns(2)
            b1.metric("Total Invoice",  f"₹{result_df['Invoice Amount'].sum():,.2f}")
            b2.metric("GT Total",       f"₹{valid['GT (As Per Calc)'].sum():,.2f}")
            b1.metric("Selling Total",  f"₹{valid['Selling Price'].sum():,.2f}")
            b2.metric("Total Received", f"₹{valid['Received Amount'].sum():,.2f}")
            net = valid["Difference"].sum()
            b1.metric(
                "Net Diff vs PWN", f"₹{net:,.2f}",
                delta=f"{'▲' if net >= 0 else '▼'} {abs(net):,.2f}",
                delta_color="normal" if net >= 0 else "inverse",
            )
            b2.metric("Orders −ve Diff", int((valid["Difference"] < 0).sum()))

        st.info(
            "ℹ️  **Brand** → from Product name  |  "
            "**Sub-Category** → Sheet 1 SKU lookup  |  "
            "**GT** → slab on Invoice Amount  |  "
            "**Received** = Selling Price − Charges − GST − TDS − TCS  |  "
            "**Difference** = Received − (Qty × PWN)"
        )
        st.markdown("---")
        st.markdown("#### 📋 Per-Order Charges Detail")
        charge_cols = [
            "Order Id","SKU","Brand Name","Sub-Category","Charge Method",
            "Invoice Amount","GT (As Per Calc)","Selling Price","Commission",
            "Collection Fee","Fixed Fee","Total Charges","GST on Charges",
            "Taxable Value","TDS","TCS","Total Deductions","Received Amount",
        ]
        avail_charge = [c for c in charge_cols if c in result_df.columns]
        st.dataframe(style_table(result_df[avail_charge]), use_container_width=True, height=480)
        st.markdown("---")
        st.markdown("#### 🧾 Grand Summary Table")
        st.dataframe(summary_df, use_container_width=True)

    # ── TAB 3: CATEGORY BREAKDOWN ──────────────────────────────────────────────
    with tab3:
        st.markdown('<div class="section-header">📊 Sub-Category-wise Breakdown</div>', unsafe_allow_html=True)
        cat_money = [c for c in cat_df.columns if c not in ("Sub-Category","Orders")]
        st.dataframe(
            style_table(cat_df, diff_col="Net Difference").format({c: "₹{:.2f}" for c in cat_money}),
            use_container_width=True,
        )
        st.markdown("---")
        st.markdown("#### 🔢 Charge Components (per Sub-Category)")
        comp_cols = [
            "Sub-Category","Orders","GT Total","Commission","Collection Fee",
            "Fixed Fee","GST Total","TDS Total","TCS Total","Total Deductions",
        ]
        comp_money = [c for c in comp_cols if c not in ("Sub-Category","Orders")]
        avail_comp = [c for c in comp_cols if c in cat_df.columns]
        st.dataframe(
            cat_df[avail_comp].style.format({c: "₹{:.2f}" for c in comp_money}),
            use_container_width=True,
        )

    # ── TAB 4: BRAND BREAKDOWN (NEW) ───────────────────────────────────────────
    with tab4:
        st.markdown('<div class="section-header">🏷️ Brand-wise Breakdown</div>', unsafe_allow_html=True)
        brand_money = [c for c in brand_df.columns if c not in ("Brand","Orders")]
        st.dataframe(
            style_table(brand_df, diff_col="Net Difference").format({c: "₹{:.2f}" for c in brand_money}),
            use_container_width=True,
        )
        st.markdown("---")
        st.markdown("#### 📋 Brand × Sub-Category Detail")
        valid2 = result_df[result_df["Received Amount"].notna()]
        bxc = (
            valid2.groupby(["Brand Name","Sub-Category"])
            .agg(
                Orders=("Order Id","count"),
                Invoice=("Invoice Amount","sum"),
                Received=("Received Amount","sum"),
                Net_Diff=("Difference","sum"),
            )
            .reset_index().round(2)
        )
        bxc.columns = ["Brand","Sub-Category","Orders","Invoice Total","Received Total","Net Difference"]
        bxc_money = ["Invoice Total","Received Total","Net Difference"]
        st.dataframe(
            bxc.style.format({c: "₹{:.2f}" for c in bxc_money}),
            use_container_width=True, height=400,
        )

else:
    # ── Welcome / instructions screen ─────────────────────────────────────────
    col_l, col_r = st.columns([3, 2])
    with col_l:
        st.markdown("""
<div style="background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:1.6rem 2rem;margin-bottom:1rem">

### 👈 Upload files in the sidebar to begin

| File | Description |
|------|-------------|
| **Order File(s)** | Flipkart Seller Hub export (CSV / XLSX / XLS) |
| **Data Excel** | Yash Gallery workbook — 3 sheets |
| **Replace SKU Excel** *(optional)* | Seller SKU → OMS SKU mapping |

</div>
""", unsafe_allow_html=True)

        st.markdown("""
<div style="background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:1.6rem 2rem;margin-bottom:1rem">

### ✅ Brand Detection

| Product starts with | Brand used |
|---------------------|------------|
| `Yash Gallery` | Yash Gallery |
| `KALINI` | KALINI |
| `Tasrika` | Tasrika |
| `AKIKO` | Pushpa |
| `HouseOfCommon` | Pushpa |
| `IKRASS` | IKRASS |

</div>
""", unsafe_allow_html=True)

    with col_r:
        st.markdown("""
<div style="background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:1.6rem 2rem">

### ✅ Calculation Logic

**Step 1** — Brand from Product name

**Step 2** — Sub-category from Sheet 1

**Step 3** — Slab lookups from Sheet 0:

| Charge | Input |
|--------|-------|
| GT | Invoice Amount |
| Commission | Selling Price |
| Collection | Selling Price |

**Step 4** — Final amounts:
```
Total Charges   = Commission + Collection + Fixed Fee
GST on Charges  = Total Charges × 18%
Taxable Value   = Selling Price − SP/105×5
TDS             = TV × 0.1%
TCS             = TV × 0.5%
Received Amount = SP − Charges − GST − TDS − TCS
Difference      = Received − (Qty × PWN)
```

</div>
""", unsafe_allow_html=True)
